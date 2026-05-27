# =============================================================================
# compare_replacement_plan.py
# -----------------------------------------------------------------------------
# Compara el Plan Maestro de Reemplazos de Ductos contra la Planilla de
# Control GIS, identificando el estado de avance por ducto:
# Entregado / Pendiente / En revisión / Rechazado.
#
# Autora : Denise Hernández — GIS Analyst
# Entorno: Python 3.x + pandas + openpyxl
# =============================================================================

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os
import datetime

# =============================================================================
# CONFIGURACIÓN
# =============================================================================

# Ruta al Plan Maestro de Reemplazos (.xlsx)
MASTER_PLAN_PATH = r"C:\Ruta\A\master_plan_reemplazos.xlsx"
MASTER_PLAN_SHEET = "Plan"
MASTER_PLAN_ID_COL = "Pipeline_ID"          # columna con el ID del ducto

# Ruta a la Planilla Control GIS (.xlsx)
GIS_CONTROL_PATH = r"C:\Ruta\A\control_gis_reemplazos.xlsx"
GIS_CONTROL_SHEET = "Control"
GIS_CONTROL_ID_COL = "Pipeline_ID"          # columna con el ID del ducto
GIS_CONTROL_STATUS_COL = "GIS_Status"       # columna con el estado en GIS

# Valores de estado en la planilla control
STATUS_DELIVERED = "Delivered"              # cargado en GIS
STATUS_IN_REVIEW = "In review"              # recibido con observaciones abiertas
STATUS_REJECTED = "Rejected"               # devuelto — pendiente corrección

# Carpeta de salida
OUTPUT_FOLDER = r"C:\Ruta\A\Salida"
OUTPUT_FILENAME = f"comparison_plan_vs_gis_{datetime.date.today().isoformat()}.xlsx"

# =============================================================================
# COLORES PARA EL EXCEL DE SALIDA
# =============================================================================

COLOR_DELIVERED = "C6EFCE"   # verde claro
COLOR_PENDING   = "FFEB9C"   # amarillo claro
COLOR_REVIEW    = "BDD7EE"   # azul claro
COLOR_REJECTED  = "FFC7CE"   # rojo claro
COLOR_HEADER    = "1E4D78"   # azul oscuro (encabezado)

# =============================================================================
# FUNCIONES
# =============================================================================

def load_excel(path, sheet, id_col):
    """Carga un Excel y valida que exista la columna ID."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"Archivo no encontrado: {path}")

    df = pd.read_excel(path, sheet_name=sheet, dtype=str)
    df.columns = df.columns.str.strip()

    if id_col not in df.columns:
        raise ValueError(
            f"Columna '{id_col}' no encontrada en {path}.\n"
            f"Columnas disponibles: {list(df.columns)}"
        )

    df[id_col] = df[id_col].str.strip().str.upper()
    return df


def classify_status(pipeline_id, gis_df, id_col, status_col):
    """
    Determina el estado de un ducto del plan maestro
    en base a si aparece (y cómo) en la planilla control GIS.
    """
    match = gis_df[gis_df[id_col] == pipeline_id]

    if match.empty:
        return "Pending"

    status = match.iloc[0].get(status_col, "").strip()

    if status == STATUS_DELIVERED:
        return "Delivered"
    elif status == STATUS_IN_REVIEW:
        return "In review"
    elif status == STATUS_REJECTED:
        return "Rejected"
    else:
        return f"Unknown ({status})"


def get_fill(status):
    """Devuelve el PatternFill correspondiente al estado."""
    mapping = {
        "Delivered":  COLOR_DELIVERED,
        "Pending":    COLOR_PENDING,
        "In review":  COLOR_REVIEW,
        "Rejected":   COLOR_REJECTED,
    }
    color = mapping.get(status, "FFFFFF")
    return PatternFill("solid", fgColor=color)


def export_comparison(results_df, output_path):
    """Exporta el DataFrame de resultados a un Excel formateado."""
    wb = openpyxl.Workbook()

    # ── Hoja 1: Detalle completo ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Plan vs GIS"

    headers = list(results_df.columns)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in results_df.iterrows():
        status = row.get("Comparison_Status", "")
        fill = get_fill(status)
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
            cell.fill = fill

    # Ajuste de anchos
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # ── Hoja 2: Resumen ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    summary_headers = ["Status", "Count", "Percentage"]
    for col_idx, h in enumerate(summary_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)

    total = len(results_df)
    counts = results_df["Comparison_Status"].value_counts()

    for row_idx, (status, count) in enumerate(counts.items(), start=2):
        pct = f"{count / total * 100:.1f}%"
        ws2.cell(row=row_idx, column=1, value=status).fill = get_fill(status)
        ws2.cell(row=row_idx, column=2, value=count)
        ws2.cell(row=row_idx, column=3, value=pct)

    ws2.cell(row=len(counts) + 2, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=len(counts) + 2, column=2, value=total).font = Font(bold=True)
    ws2.cell(row=len(counts) + 2, column=3, value="100%").font = Font(bold=True)

    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 18

    wb.save(output_path)
    print(f"\n[OK] Reporte exportado: {output_path}")


# =============================================================================
# EJECUCIÓN PRINCIPAL
# =============================================================================

def main():
    print("=" * 60)
    print("  Pipeline Replacement — Plan vs GIS Comparison")
    print(f"  {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # Cargar archivos
    print("\n[1/4] Cargando Plan Maestro...")
    master_df = load_excel(MASTER_PLAN_PATH, MASTER_PLAN_SHEET, MASTER_PLAN_ID_COL)
    print(f"      {len(master_df)} ductos en el plan maestro.")

    print("[2/4] Cargando Planilla Control GIS...")
    gis_df = load_excel(GIS_CONTROL_PATH, GIS_CONTROL_SHEET, GIS_CONTROL_ID_COL)
    print(f"      {len(gis_df)} registros en la planilla control.")

    # Clasificar cada ducto del plan
    print("[3/4] Comparando...")
    master_df["Comparison_Status"] = master_df[MASTER_PLAN_ID_COL].apply(
        lambda pid: classify_status(pid, gis_df, GIS_CONTROL_ID_COL, GIS_CONTROL_STATUS_COL)
    )

    # Resumen en consola
    print("\n--- RESUMEN ---")
    total = len(master_df)
    counts = master_df["Comparison_Status"].value_counts()
    for status, count in counts.items():
        pct = count / total * 100
        bar = "█" * int(pct / 5)
        print(f"  {status:<15} {count:>4}  ({pct:5.1f}%)  {bar}")
    print(f"  {'TOTAL':<15} {total:>4}")

    # Exportar
    print("\n[4/4] Exportando reporte...")
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    output_path = os.path.join(OUTPUT_FOLDER, OUTPUT_FILENAME)
    export_comparison(master_df, output_path)


if __name__ == "__main__":
    main()
